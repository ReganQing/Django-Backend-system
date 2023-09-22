#!usr/env/python3
# -*- coding: utf-8 -*-

"""
# File       : department.py
# Time       : 2023/2/28 13:51
# Author     : Ron
# version    : python 3.9
# software   : Pycharm
# Description：部门信息管理
"""

from django.shortcuts import render, redirect

import web.utils.form
from web import models
from web.utils.pagination import Pagination
from web.utils.form import DepartmentModelForm

# Create your views here.


def depart_list(request):
    """部门列表"""

    # 去数据库中获取所有部门列表
    depart_info = models.Department.objects.all()
    depart_obj = Pagination(request, depart_info, page_size=4)
    context = {
        "query_set": depart_obj.page_queryset,
        "page_string": depart_obj.html(),
    }
    return render(request, "depart_list.html", context)


def depart_add(request):
    """添加部门"""

    if request.method == "GET":
        form = DepartmentModelForm()
        return render(request, "depart_add.html", {"form": form})

    # 用户POST进行数据校验
    form = DepartmentModelForm(data=request.POST)
    if form.is_valid():
        # 将数据保存到用户信息表中
        form.save()
        return redirect("/depart/list/")

    # 若校验失败(在页面上显示错误信息)

    return render(request, "depart_add.html", {"form": form})


def depart_del(request, nid):
    """删除功能"""
    # http://127.0.0.1:8000/depart/del/?nid=1
    models.Department.objects.filter(id=nid).delete()

    return redirect("/depart/list/")


def depart_edit(request, nid):
    """修改功能"""
    row_object = models.Department.objects.filter(id=nid).first()

    if request.method == "GET":
        # 根据ID去数据库获取要编辑的那一行数据(对象)
        form = DepartmentModelForm(instance=row_object)
        return render(request, "depart_edit.html", {"form": form})

    form = DepartmentModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        # 默认保存的是用户输入的所有数据，如果想要在用户输入以外增加一点值
        # form.instance.字段名 = 值
        form.save()
        return redirect("/depart/list/")
    return render(request, "depart_edit.html", {"form": form})


def depart_multi(request):
    """批量上传"""
    from openpyxl import load_workbook

    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")
    # print(type(file_object))

    # 2.对象传递给openpyxl，由openpyxl读取文件内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(department_name=text).exists()
        if not exists:
            models.Department.objects.create(department_name=text)

    return redirect("/depart/list/")
